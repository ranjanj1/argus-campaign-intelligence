from __future__ import annotations

from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

from scripts.models.client_context import ClientProfile, CampaignSeed

# Color constants
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD_FONT = Font(bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_TO_IDX = {m: i + 1 for i, m in enumerate(MONTHS)}


def _header_row(ws, headers: list[str], row: int = 1) -> None:
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def generate_budget_allocation_xlsx(profile: ClientProfile) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    _build_quarterly_breakdown(wb, profile)
    _build_campaign_budget(wb, profile)
    _build_forecast_vs_actual(wb, profile)
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


# ---------------------------------------------------------------------------
# Sheet 1: quarterly_breakdown — months × channels
# ---------------------------------------------------------------------------

def _build_quarterly_breakdown(wb: openpyxl.Workbook, profile: ClientProfile) -> None:
    ws = wb.create_sheet("quarterly_breakdown")

    # Aggregate spend by (month_name, channel)
    spend_grid: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    channels_seen: set[str] = set()

    for camp in profile.campaigns:
        start = camp.start_date
        end = camp.end_date
        month_name = MONTHS[start.month - 1]
        spend_grid[month_name][camp.channel] += camp.spend
        channels_seen.add(camp.channel)

    channels = sorted(channels_seen)
    headers = ["Month"] + channels + ["Total"]
    _header_row(ws, headers)

    # Auto column width
    ws.column_dimensions["A"].width = 12
    for i in range(len(channels) + 1):
        ws.column_dimensions[get_column_letter(i + 2)].width = 16

    for row_idx, month in enumerate(MONTHS, 2):
        ws.cell(row=row_idx, column=1, value=month).font = BOLD_FONT
        total = 0.0
        for col_idx, ch in enumerate(channels, 2):
            val = round(spend_grid[month].get(ch, 0.0), 2)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.number_format = '"$"#,##0.00'
            cell.border = THIN_BORDER
            total += val
        # Total column
        total_cell = ws.cell(row=row_idx, column=len(channels) + 2, value=round(total, 2))
        total_cell.number_format = '"$"#,##0.00'
        total_cell.font = BOLD_FONT
        total_cell.border = THIN_BORDER

    # Grand total row
    total_row = len(MONTHS) + 2
    ws.cell(row=total_row, column=1, value="Total").font = BOLD_FONT
    grand_total = 0.0
    for col_idx, ch in enumerate(channels, 2):
        col_total = sum(spend_grid[m].get(ch, 0.0) for m in MONTHS)
        cell = ws.cell(row=total_row, column=col_idx, value=round(col_total, 2))
        cell.number_format = '"$"#,##0.00'
        cell.font = BOLD_FONT
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        grand_total += col_total
    gt_cell = ws.cell(row=total_row, column=len(channels) + 2, value=round(grand_total, 2))
    gt_cell.number_format = '"$"#,##0.00'
    gt_cell.font = Font(bold=True, color="FFFFFF")
    gt_cell.fill = HEADER_FILL


# ---------------------------------------------------------------------------
# Sheet 2: campaign_budget — per-campaign allocated vs actual
# ---------------------------------------------------------------------------

def _build_campaign_budget(wb: openpyxl.Workbook, profile: ClientProfile) -> None:
    ws = wb.create_sheet("campaign_budget")

    headers = [
        "campaign_id", "campaign_name", "channel", "status",
        "allocated_budget", "actual_spend", "variance", "variance_pct",
    ]
    _header_row(ws, headers)

    col_widths = [14, 45, 18, 12, 18, 16, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row_idx, camp in enumerate(profile.campaigns, 2):
        variance = round(camp.spend - camp.total_budget, 2)
        variance_pct = round(variance / camp.total_budget * 100, 1) if camp.total_budget else 0.0

        row_data = [
            camp.campaign_id,
            camp.name,
            camp.channel,
            camp.status,
            camp.total_budget,
            camp.spend,
            variance,
            variance_pct,
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (5, 6, 7):
                cell.number_format = '"$"#,##0.00'
            if col_idx == 8:
                cell.number_format = '0.0"%"'

        # Conditional fill on variance column
        var_cell = ws.cell(row=row_idx, column=7)
        if variance < 0:
            var_cell.fill = GREEN_FILL   # under budget = good
        elif variance > camp.total_budget * 0.05:
            var_cell.fill = RED_FILL     # >5% over budget = bad

    # Totals row
    n = len(profile.campaigns) + 2
    ws.cell(row=n, column=1, value="TOTAL").font = BOLD_FONT
    total_budget = sum(c.total_budget for c in profile.campaigns)
    total_spend = sum(c.spend for c in profile.campaigns)
    total_var = total_spend - total_budget
    total_var_pct = round(total_var / total_budget * 100, 1) if total_budget else 0.0

    for col_idx, val in zip([5, 6, 7, 8], [total_budget, total_spend, total_var, total_var_pct]):
        cell = ws.cell(row=n, column=col_idx, value=round(val, 2))
        cell.font = BOLD_FONT
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        if col_idx in (5, 6, 7):
            cell.number_format = '"$"#,##0.00'


# ---------------------------------------------------------------------------
# Sheet 3: forecast_vs_actual — quarterly view
# ---------------------------------------------------------------------------

def _build_forecast_vs_actual(wb: openpyxl.Workbook, profile: ClientProfile) -> None:
    ws = wb.create_sheet("forecast_vs_actual")

    headers = [
        "quarter", "channel", "forecast_spend", "actual_spend",
        "variance", "variance_pct", "campaigns_count",
    ]
    _header_row(ws, headers)

    col_widths = [10, 18, 16, 16, 14, 14, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Aggregate by quarter and channel
    def quarter_of(d) -> str:
        return f"Q{(d.month - 1) // 3 + 1} {d.year}"

    agg: dict[tuple[str, str], dict] = defaultdict(
        lambda: {"forecast": 0.0, "actual": 0.0, "count": 0}
    )
    for camp in profile.campaigns:
        q = quarter_of(camp.start_date)
        key = (q, camp.channel)
        agg[key]["forecast"] += camp.total_budget
        agg[key]["actual"] += camp.spend
        agg[key]["count"] += 1

    row_idx = 2
    for (quarter, channel), vals in sorted(agg.items()):
        forecast = round(vals["forecast"], 2)
        actual = round(vals["actual"], 2)
        variance = round(actual - forecast, 2)
        var_pct = round(variance / forecast * 100, 1) if forecast else 0.0

        row_data = [quarter, channel, forecast, actual, variance, var_pct, vals["count"]]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            if col_idx in (3, 4, 5):
                cell.number_format = '"$"#,##0.00'
            if col_idx == 6:
                cell.number_format = '0.0"%"'

        # Conditional formatting on variance
        var_cell = ws.cell(row=row_idx, column=5)
        if variance < 0:
            var_cell.fill = GREEN_FILL
        elif variance > forecast * 0.1:
            var_cell.fill = RED_FILL

        row_idx += 1
